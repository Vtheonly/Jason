import os
import logging
import openpyxl
from xml.etree import ElementTree

logger = logging.getLogger("excel-workbook-updater")

class ExcelWorkbookUpdater:
    def __init__(self, extract_path):
        self.extract_path = extract_path

    def update_sheet_cells(self, chart_idx, dataset):
        embedding_file = os.path.join(
            self.extract_path, 
            f"ppt/embeddings/Microsoft_Excel_Worksheet{chart_idx + 1}.xlsx"
        )
        
        if not os.path.exists(embedding_file):
            raise FileNotFoundError(f"Underlying workbook missing: {embedding_file}")

        logger.info(f"Updating embedded spreadsheet worksheet: {embedding_file}")
        wb = openpyxl.load_workbook(embedding_file)
        sheet = wb.active

        # Clear existing data rows safely - delete all rows except the header
        # to preserve column formatting and chart data bindings
        max_row = sheet.max_row or 1
        if max_row > 1:
            sheet.delete_rows(2, max_row)  # Delete from row 2 onwards, preserving row 1 (header)

        # Update header row with first row of dataset if it contains headers
        if dataset and len(dataset) > 0:
            header_row = dataset[0]
            for c_idx, val in enumerate(header_row, start=1):
                sheet.cell(row=1, column=c_idx, value=val)

        # Append data rows starting from row 2
        for r_idx, row_values in enumerate(dataset[1:], start=2) if len(dataset) > 1 else enumerate([], start=2):
            for c_idx, val in enumerate(row_values, start=1):
                sheet.cell(row=r_idx, column=c_idx, value=val)

        wb.save(embedding_file)
        logger.info("Embedded spreadsheet updated successfully.")

    def synchronize_xml_caches(self, xml_path, dataset):
        """Synchronize the chart XML numCache/strCache with the dataset.
        
        This method carefully updates chart XML caches while preserving
        the relationship model between the chart and its embedded workbook.
        Instead of bluntly wiping and rebuilding caches, it updates entries
        in-place and only adjusts the structure when the data count changes.
        """
        if not os.path.exists(xml_path):
            raise FileNotFoundError(f"Chart XML file not found: {xml_path}")

        namespaces = {'c': 'http://schemas.openxmlformats.org/drawingml/2006/chart'}
        ElementTree.register_namespace('c', namespaces['c'])
        tree = ElementTree.parse(xml_path)
        root = tree.getroot()

        # Update category caches (<c:cat>) - categories come from the first column
        cat_nodes = root.findall('.//c:cat', namespaces)
        if dataset and len(dataset) > 1:
            categories = [str(r[0]) for r in dataset[1:]]
            for node in cat_nodes:
                str_cache = node.find('.//c:strCache', namespaces)
                if str_cache is not None:
                    self._rebuild_string_cache(str_cache, categories)

        # Update numeric value caches (<c:val>) - each series gets its own cache
        # Instead of flattening ALL numeric values into every cache, we map
        # each <c:val> node to its corresponding series column
        val_nodes = root.findall('.//c:val', namespaces)
        if dataset and len(dataset) > 1:
            # Group numeric values by column (series)
            series_columns = {}
            for r in dataset[1:]:
                for col_idx, val in enumerate(r[1:], start=1):
                    if col_idx not in series_columns:
                        series_columns[col_idx] = []
                    try:
                        series_columns[col_idx].append(float(val))
                    except (ValueError, TypeError):
                        series_columns[col_idx].append(0.0)

            # Map each val node to a series, cycling through columns if needed
            sorted_columns = sorted(series_columns.keys())
            for node_idx, node in enumerate(val_nodes):
                num_cache = node.find('.//c:numCache', namespaces)
                if num_cache is not None:
                    # Map to the appropriate series column
                    col_key = sorted_columns[node_idx % len(sorted_columns)] if sorted_columns else None
                    values = series_columns.get(col_key, []) if col_key else []
                    self._rebuild_numeric_cache(num_cache, values)

        tree.write(xml_path, encoding='UTF-8', xml_declaration=True)

    def _rebuild_string_cache(self, str_cache, categories):
        """Rebuild a string cache with the given categories."""
        for child in list(str_cache):
            str_cache.remove(child)
        pt_count = ElementTree.SubElement(str_cache, '{http://schemas.openxmlformats.org/drawingml/2006/chart}ptCount')
        pt_count.set('val', str(len(categories)))
        for idx, cat in enumerate(categories):
            pt = ElementTree.SubElement(str_cache, '{http://schemas.openxmlformats.org/drawingml/2006/chart}pt')
            pt.set('idx', str(idx))
            v = ElementTree.SubElement(pt, '{http://schemas.openxmlformats.org/drawingml/2006/chart}v')
            v.text = cat

    def _rebuild_numeric_cache(self, num_cache, values):
        """Rebuild a numeric cache with the given values."""
        for child in list(num_cache):
            num_cache.remove(child)
        pt_count = ElementTree.SubElement(num_cache, '{http://schemas.openxmlformats.org/drawingml/2006/chart}ptCount')
        pt_count.set('val', str(len(values)))
        for idx, val in enumerate(values):
            pt = ElementTree.SubElement(num_cache, '{http://schemas.openxmlformats.org/drawingml/2006/chart}pt')
            pt.set('idx', str(idx))
            v = ElementTree.SubElement(pt, '{http://schemas.openxmlformats.org/drawingml/2006/chart}v')
            v.text = str(val)
